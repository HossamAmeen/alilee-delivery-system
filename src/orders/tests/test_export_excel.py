import io

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status

from orders.models import OrderStatus


@pytest.mark.django_db
class TestOrderExportExcel:
    def setup_method(self):
        self.url = reverse("orders-export-excel")

    def test_export_excel_two_sheets_success(
        self, admin_client, assigned_order, created_order
    ):

        # 2. Prepare an Un-financial Order (created_order - it is in CREATED status)
        url = reverse("orders-detail", kwargs={"pk": assigned_order.id})

        update_payload = {"status": OrderStatus.DELIVERED}

        response = admin_client.patch(url, data=update_payload, format="json")
        assert response.status_code == status.HTTP_200_OK

        # 3. Call the export endpoint
        response = admin_client.get(self.url)

        # 4. Assertions
        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Load the workbook from the response content
        wb = load_workbook(io.BytesIO(response.content))

        # Verify sheet names
        assert "Financial Orders" in wb.sheetnames
        assert "Un-financial Orders" in wb.sheetnames

        # Verify content of "Financial Orders" sheet
        ws1 = wb["Financial Orders"]
        # Header is row 1. Data should be in row 2.
        # Check if the assigned_order tracking number is in the second row
        found_financial = False
        for row in range(2, ws1.max_row + 1):
            if str(ws1.cell(row=row, column=4).value) == str(
                assigned_order.tracking_number
            ):
                found_financial = True
                break
        assert (
            found_financial
        ), "Assigned order (financial) not found in 'Financial Orders' sheet"

        # Verify content of "Un-financial Orders" sheet
        ws2 = wb["Un-financial Orders"]
        found_un_financial = False
        for row in range(2, ws2.max_row + 1):
            if str(ws2.cell(row=row, column=3).value) == str(
                created_order.tracking_number
            ):
                found_un_financial = True
                break
        assert (
            found_un_financial
        ), "Created order (un-financial) not found in 'Un-financial Orders' sheet"
