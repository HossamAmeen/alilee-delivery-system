from datetime import date, datetime, timedelta
from io import BytesIO

from django.core.exceptions import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter

from orders.models import Order, OrderStatus, ProductPaymentStatus
from users.models import Trader
from utilities.exceptions import CustomValidationError


class DeliveryAssignmentService:
    @staticmethod
    def assign_driver(order, driver):
        forbidden_statuses = [
            OrderStatus.DELIVERED,
            OrderStatus.CANCELLED,
            OrderStatus.ASSIGNED,
        ]
        if order.driver:
            raise CustomValidationError("Order is already assigned to a driver.")
        if order.status in forbidden_statuses:
            raise CustomValidationError("Order cannot be assigned due to its status.")
        order.driver = driver
        order.status = OrderStatus.ASSIGNED
        order.save()

        return order


class OrderExportService:
    @staticmethod
    def get_export_queryset(params):
        queryset = Order.objects.select_related(
            "driver", "trader", "customer", "delivery_zone"
        ).order_by("-id")

        trader_id = params.get("trader")
        tracking_numbers = params.get("tracking_numbers")
        reference_codes = params.get("reference_codes")
        status = params.get("status")
        driver_id = params.get("driver")
        today = date.today()
        date_from = params.get("date_from")
        file_name_suffix = ""

        if not date_from:
            date_from = today - timedelta(days=7)
            file_name_suffix += f"{date_from.strftime('%Y-%m-%d')}"
        else:
            try:
                date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            except ValueError:
                raise CustomValidationError(
                    message="Invalid date format. Use YYYY-MM-DD."
                )

        date_to = params.get("date_to")
        if not date_to:
            date_to = today
            file_name_suffix += f"_{date_to.strftime('%Y-%m-%d')}"
        else:
            try:
                date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
                file_name_suffix += f"_{date_to.strftime('%Y-%m-%d')}"
            except ValueError:
                raise CustomValidationError(
                    message="Invalid date format. Use YYYY-MM-DD."
                )

        if date_from:
            try:
                queryset = queryset.filter(status_changed_at__date__gte=date_from)
            except (ValueError, ValidationError):
                raise CustomValidationError(
                    message="Invalid date format. Use YYYY-MM-DD."
                )

        if date_to:
            if date_from:
                if date_from > date_to:
                    raise CustomValidationError(
                        message="Date from cannot be greater than date to."
                    )

                try:
                    if (date_to - date_from).days > 7:
                        pass
                except ValueError:
                    raise CustomValidationError(
                        message="Invalid date format. Use YYYY-MM-DD."
                    )

            try:
                queryset = queryset.filter(status_changed_at__date__lte=date_to)
            except (ValueError, ValidationError):
                raise CustomValidationError(
                    message="Invalid date format. Use YYYY-MM-DD."
                )

        if trader_id:
            queryset = queryset.filter(trader_id=trader_id)
            trader = Trader.objects.get(id=trader_id)
            file_name_suffix += f"_{trader.full_name}"

        if tracking_numbers:
            tracking_numbers_list = [tn.strip() for tn in tracking_numbers.split(",")]
            queryset = queryset.filter(tracking_number__in=tracking_numbers_list)

        if reference_codes:
            reference_codes_list = [rc.strip() for rc in reference_codes.split(",")]
            queryset = queryset.filter(reference_code__in=reference_codes_list)

        if status:
            status_list = [s.strip() for s in status.split(",")]
            queryset = queryset.filter(status__in=status_list)
            file_name_suffix += f"_status_{status}"

        if driver_id:
            queryset = queryset.filter(driver_id=driver_id)

        if queryset.count() > 5000:
            raise CustomValidationError(
                message="Cannot export more than 5000 orders at once."
            )

        return queryset, f"orders_{file_name_suffix}"

    @staticmethod
    def _calculate_order_financials(order):
        trader_cost = (
            order.trader_cost if order.trader_cost else order.trader_merchant_cost
        )
        trader_commission = 0
        office = 0

        if order.product_payment_status == ProductPaymentStatus.COD:
            if order.status == OrderStatus.DELIVERED:
                if order.product_cost > trader_cost:
                    trader_commission = order.product_cost - trader_cost
                else:
                    office = trader_cost - order.product_cost
        if order.product_payment_status == ProductPaymentStatus.PAID:
            office = trader_cost

        if order.status == OrderStatus.CREATED or order.status == OrderStatus.ASSIGNED:
            trader_commission = 0
            office = 0

        return trader_cost, trader_commission, office

    @classmethod
    def generate_csv(cls, queryset, writer):
        writer.writerow(
            [
                "تاريخ الاضافة",
                "رقم التتبع",
                "رمز المرجع",
                "اسم التاجر",
                "العنوان",
                "اسم المستلم",
                "رقم هاتف المستلم"
                "الحالة",
                "سعر الشحنه",
                "حالة الدفع",
                "رسوم الشحن",
                "فلوس المكتب",
                "فلوس التاجر",
                "فرق الفلوس",
            ]
        )

        total_trader_commission = 0
        total_office = 0
        for order in queryset:
            trader_cost, trader_commission, office = cls._calculate_order_financials(
                order
            )

            total_trader_commission += trader_commission
            total_office += office
            writer.writerow(
                [
                    order.created.strftime("%Y-%m-%d"),
                    str(order.tracking_number),
                    str(order.reference_code),
                    str(order.trader.full_name if order.trader else ""),
                    str(order.delivery_zone.name if order.delivery_zone else ""),
                    str(order.customer.name if order.customer else ""),
                    str(order.customer.phone if order.customer else ""),
                    str(order.status_ar),
                    str(order.product_cost),
                    str(order.product_payment_status_ar),
                    str(trader_cost),
                    str(office),
                    str(trader_commission),
                    str(office - trader_commission),
                ]
            )

        writer.writerow(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                total_office,
                total_trader_commission,
                total_office - total_trader_commission,
                "اجمالي",
            ]
        )

    @classmethod
    def generate_excel(cls, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        ws.append(
            [
                "تاريخ الاضافة",
                "تاريخ التغيير",
                "رقم التتبع",
                "رمز المرجع",
                "اسم التاجر",
                "العنوان",
                "اسم المستلم",
                "رقم هاتف المستلم",
                "الحالة",
                "سعر الشحنه",
                "حالة الدفع",
                "رسوم الشحن",
                "فلوس المكتب",
                "فلوس التاجر",
                "فرق الفلوس",
            ]
        )

        total_trader_commission = 0
        total_office = 0

        # Create a mapping of status to color (column F is the status column)
        for row_idx, order in enumerate(queryset, start=2):  # Start from row 2 (1-based) to skip header
            trader_cost, trader_commission, office = cls._calculate_order_financials(
                order
            )

            total_trader_commission += trader_commission
            total_office += office
            customer_name = ""
            customer_phone = ""
            if order.customer:
                customer_name = order.customer.name
                customer_phone = order.customer.phone                
            ws.append(
                [
                    order.created.strftime("%Y-%m-%d"),
                    order.status_changed_at.strftime("%Y-%m-%d") if order.status_changed_at else "",
                    str(order.tracking_number),
                    str(order.reference_code),
                    str(order.trader.full_name if order.trader else ""),
                    str(order.delivery_zone.name if order.delivery_zone else ""),
                    str(customer_name),
                    str(customer_phone),
                    str(order.status_ar),
                    str(order.product_cost),
                    str(order.product_payment_status_ar),
                    str(trader_cost),
                    str(office),
                    str(trader_commission),
                    str(office - trader_commission),
                ]
            )
            
            # Get the status cell (column H)
            status_cell = ws[f'H{row_idx}']
            # Apply the color based on status
            status_cell.font = Font(color=order.status_color.lstrip('#'))  # Remove '#' if present

        ws.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                total_office,
                total_trader_commission,
                total_office - total_trader_commission,
                "اجمالي",
            ]
        )

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 2

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
